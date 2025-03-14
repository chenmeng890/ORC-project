import openpyxl
import json
import os
from datetime import datetime

class ExcelProcessor:
    def __init__(self):
        self.workbook = None
        self.worksheet = None
        self.invoice_count = 0
        self.callback = None  # 添加回调函数

    def create_new_workbook(self):
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = '发票识别结果'
        self.invoice_count = 0  # 重置计数器
        
        # 设置表头
        headers = ['文件名', '识别时间', '发票类型', '发票代码', '发票号码', '开票日期', 
                  '购买方名称', '销售方名称', '金额', '税率', '税额', '价税合计', '备注']
        for col, header in enumerate(headers, 1):
            self.worksheet.cell(row=1, column=col, value=header)

    def set_callback(self, callback):
        self.callback = callback

    def notify_progress(self, filename, status, message=''):
        if self.callback:
            self.callback(filename, status, message)

    def add_result(self, filename, ocr_result):
        if not self.workbook:
            self.create_new_workbook()

        try:
            # 获取下一个空行
            next_row = self.worksheet.max_row + 1

            # 添加基本信息
            self.worksheet.cell(row=next_row, column=1, value=filename)
            self.worksheet.cell(row=next_row, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

            # 解析OCR结果
            invoice_info = self.extract_invoice_info(ocr_result)
            
            # 填充发票信息
            self.worksheet.cell(row=next_row, column=3, value=invoice_info.get('发票类型', ''))
            self.worksheet.cell(row=next_row, column=4, value=invoice_info.get('发票代码', ''))
            self.worksheet.cell(row=next_row, column=5, value=invoice_info.get('发票号码', ''))
            self.worksheet.cell(row=next_row, column=6, value=invoice_info.get('日期', ''))
            self.worksheet.cell(row=next_row, column=7, value=invoice_info.get('购买方名称', ''))
            self.worksheet.cell(row=next_row, column=8, value=invoice_info.get('销售方名称', ''))
            self.worksheet.cell(row=next_row, column=9, value=invoice_info.get('金额', ''))
            self.worksheet.cell(row=next_row, column=10, value=invoice_info.get('税率', ''))
            self.worksheet.cell(row=next_row, column=11, value=invoice_info.get('税额', ''))
            self.worksheet.cell(row=next_row, column=12, value=invoice_info.get('价税合计', ''))
            self.worksheet.cell(row=next_row, column=13, value=invoice_info.get('备注', ''))
            
            self.invoice_count += 1
            self.notify_progress(filename, 'success', '处理成功')
        except Exception as e:
            self.notify_progress(filename, 'error', f'处理失败：{str(e)}')
            raise e

    def save_workbook(self, filepath):
        if self.workbook:
            # 从filepath中获取目录和基本文件名
            directory = os.path.dirname(filepath)
            filename = os.path.basename(filepath)
            name, ext = os.path.splitext(filename)
            
            # 生成带有时间戳和发票数量的新文件名
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            new_filename = f'{name}_{timestamp}_共{self.invoice_count}张{ext}'
            new_filepath = os.path.join(directory, new_filename)
            
            self.workbook.save(new_filepath)

    def extract_invoice_info(self, ocr_result):
        # 打印原始 OCR 接口返回结果
        print("百度 OCR 接口返回结果：")
        print(json.dumps(ocr_result, ensure_ascii=False, indent=2))

        invoice_info = {
            '发票类型': '',
            '发票代码': '',
            '发票号码': '',
            '日期': '',
            '购买方名称': '',
            '销售方名称': '',
            '金额': '',
            '税率': '',
            '税额': '',
            '价税合计': '',
            '备注': ''
        }

        if not ocr_result or len(ocr_result) == 0:
            return invoice_info

        try:
            result = ocr_result[0]  # 获取第一个结果
            
            # 直接映射字段
            field_mapping = {
                'InvoiceType': '发票类型',
                'InvoiceCode': '发票代码',
                'InvoiceNum': '发票号码',
                'InvoiceDate': '日期',
                'PurchaserName': '购买方名称',
                'SellerName': '销售方名称',
                'TotalAmount': '金额',
                'AmountInFiguers': '价税合计',
                'Remarks': '备注'
            }

            for api_field, local_field in field_mapping.items():
                if api_field in result:
                    # 处理特殊字段的格式化
                    value = result[api_field]
                    if api_field in ['TotalAmount', 'AmountInFiguers']:
                        try:
                            # 确保金额为数字格式，保留2位小数
                            value = '{:.2f}'.format(float(str(value).replace(',', '')))
                        except (ValueError, TypeError) as e:
                            print(f"处理金额字段 {api_field} 时出错: {str(e)}")
                            value = ''
                    invoice_info[local_field] = value

            # 处理税率
            if 'CommodityTaxRate' in result and len(result['CommodityTaxRate']) > 0:
                tax_rate = result['CommodityTaxRate'][0].get('word', '')
                # 统一税率格式（去除百分号和空格，保留数字）
                tax_rate = tax_rate.replace('%', '').replace('％', '').strip()
                try:
                    tax_rate = '{:.0f}%'.format(float(tax_rate))
                    invoice_info['税率'] = tax_rate
                except (ValueError, TypeError) as e:
                    print(f"处理税率时出错: {str(e)}")
                    invoice_info['税率'] = ''
            
            # 处理税额
            if 'TotalTax' in result:
                try:
                    # 确保税额为数字格式，保留2位小数
                    tax_amount = '{:.2f}'.format(float(str(result['TotalTax']).replace(',', '')))
                    invoice_info['税额'] = tax_amount
                except (ValueError, TypeError) as e:
                    print(f"处理税额时出错: {str(e)}")
                    invoice_info['税额'] = ''

        except Exception as e:
            print(f"处理发票信息时出错: {str(e)}")
            # 保持默认的空值

        return invoice_info